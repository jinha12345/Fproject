from openpyxl.utils.cell import column_index_from_string

def getby(workbook, location, model, cloth_type, is_stock):
    # DB_stock_info를 기반으로 location과 type을 통해 (시작열, 끝열)을 알아내기
    DB_stock_info = {
        'NC불광': {'btm': ('PZ', 'QF'), 'top': ('LV', 'LZ'), 'acc': ('LC', 'LG')},
        'MD구리': {'btm': ('NV', 'OB'), 'top': ('KF', 'KJ'), 'acc': ('JM', 'JQ')},
        'TO분당': {'btm': ('OD', 'OJ'), 'top': ('KL', 'KP'), 'acc': ('JS', 'JW')},
        'M춘천': {'btm': ('PB', 'PH'), 'top': ('LD', 'LH'), 'acc': ('KK', 'KO')},
        'MD부평': {'btm': ('OL', 'OR'), 'top': ('KR', 'KV'), 'acc': ('JY', 'KC')},
        'NC고잔': {'btm': ('OT', 'OZ'), 'top': ('KX', 'LB'), 'acc': ('KE', 'KI')},
        'NC송파': {'btm': ('PJ', 'PP'), 'top': ('LJ', 'LN'), 'acc': ('KQ', 'KU')},
        'MD천안': {'btm': ('PR', 'PX'), 'top': ('LP', 'LT'), 'acc': ('KW', 'LA')},
        '창고'  : {'btm': ('NN', 'NT'), 'top': ('JZ', 'KD'), 'acc': ('JG', 'JK')}
    }
    
    DB_sell_info = {
        'NC불광': {'btm': ('NE', 'NK'), 'top': ('JS', 'JW'), 'acc': ('IZ', 'JD')},
        'MD구리': {'btm': ('LA', 'LG'), 'top': ('IC', 'IG'), 'acc': ('HJ', 'HN')},
        'TO분당': {'btm': ('LI', 'LO'), 'top': ('II', 'IM'), 'acc': ('HP', 'HT')},
        'M춘천': {'btm': ('MG', 'MM'), 'top': ('JA', 'JE'), 'acc': ('IH', 'IL')},
        'MD부평': {'btm': ('LQ', 'LW'), 'top': ('IO', 'IS'), 'acc': ('HV', 'HZ')},
        'NC고잔': {'btm': ('LY', 'ME'), 'top': ('IU', 'IY'), 'acc': ('IB', 'IF')},
        'NC송파': {'btm': ('MO', 'MU'), 'top': ('JG', 'JK'), 'acc': ('IN', 'IR')},
        'MD천안': {'btm': ('MW', 'NC'), 'top': ('JM', 'JQ'), 'acc': ('IT', 'IX')}
    }

    find_check = False

    if is_stock: #재고 조회
        start_col, end_col = DB_stock_info[location][cloth_type[-3:]]# if location in DB_stock_info and cloth_type[-3:] in DB_stock_info[location] else ('A', 'C')
        # 해당 시트만 가져오기
        sheet = workbook[cloth_type[-3:]]
        
        # 엑셀 시트에서 'model'을 찾아 해당 행 번호를 찾기
        for row_num, cell in enumerate(sheet['D'], start=1):
            if cell.value == model:
                find_check = True
                break
        if not find_check:
            return 'Cannotfind'
        # 해당 행과 열 범위의 값을 리스트로 가져오기
        stock_values = [sheet.cell(row_num, col).value for col in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1)]

        # 모든 값이 'None'인지 확인
        if all(value is None for value in stock_values):
            return None
        else:
            #stock_values = [0 if value is None else value for value in stock_values]
            while len(stock_values) < 7:
                stock_values.append(None)
            sum1 = sum(val if val is not None else 0 for val in stock_values) if any(val is not None for val in stock_values) else None
            stock_values.append(sum1)
            return stock_values
        


    else: #판매 조회
        start_col, end_col = DB_sell_info[location][cloth_type[-3:]]# if location in DB_sell_info and cloth_type[-3:] in DB_sell_info[location] else ('A', 'C')
        
        # 해당 시트만 가져오기
        sheet = workbook[cloth_type[-3:]]
        
        # 엑셀 시트에서 'model'을 찾아 해당 행 번호를 찾기
        for row_num, cell in enumerate(sheet['D'], start=1):
            if cell.value == model:
                break

        # 해당 행과 열 범위의 값을 리스트로 가져오기
        stock_values = [sheet.cell(row_num, col).value for col in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1)]

        # 모든 값이 'None'인지 확인
        if all(value is None for value in stock_values):
            return None
        else:
            #stock_values = [0 if value is None else value for value in stock_values]
            while len(stock_values) < 7:
                stock_values.append(None)
            sum1 = sum(val if val is not None else 0 for val in stock_values) if any(val is not None for val in stock_values) else None
            stock_values.append(sum1)
            return stock_values
        

def getSeason(workbook, model, cloth_type):
    sheet = workbook[cloth_type[-3:]]
    
    for row_num, cell in enumerate(sheet['D'], start=1):
        if cell.value == model:
            break

    return sheet.cell(row_num, 1).value

def getMSRP(workbook, model, cloth_type):
    sheet = workbook[cloth_type[-3:]]
    
    for col_num, cell in enumerate(sheet.iter_cols(min_row=5, max_row=5, values_only=True), start=1):
    # sheet.iter_cols를 사용하여 지정된 행의 각 열에서 데이터를 가져옵니다.
        if cell[0] == 'MSRP':
            break
    
    for row_num, cell in enumerate(sheet['D'], start=1):
        if cell.value == model:
            break

    return sheet.cell(row_num, col_num).value
        

def isThere(workbook, model, cloth_type):
    sheet = workbook[cloth_type[-3:]]
    
    # 엑셀 시트에서 'model'을 찾아 해당 행 번호를 찾기
    for row_num, cell in enumerate(sheet['D'], start=1):
        if cell.value == model:
            return True
    return False