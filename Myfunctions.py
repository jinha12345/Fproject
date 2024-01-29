import openpyxl

def format_price(input_number):
    # 숫자를 문자열로 변환하고 뒤집기
    reversed_str = str(input_number)[::-1]

    # 세 자리마다 쉼표 추가
    formatted_str = ",".join([reversed_str[i:i+3] for i in range(0, len(reversed_str), 3)])

    # 다시 뒤집기
    result_str = formatted_str[::-1]

    # ₩ 기호와 함께 반환
    return f"₩{result_str}"

def load_specific_sheets(file_path, sheet_names, data_only=True):
    # 빈 워크북 생성
    new_workbook = openpyxl.Workbook()

    # 기존 워크북에서 각 시트를 가져와서 새로운 워크북에 추가
    original_workbook = openpyxl.load_workbook(file_path, data_only=data_only)
    for sheet_name in sheet_names:
        original_sheet = original_workbook[sheet_name]
        new_sheet = new_workbook.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows():
            new_sheet.append([cell.value for cell in row])

    return new_workbook